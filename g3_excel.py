import openpyxl
import os

OUTPUT_FOLDER = "output"
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "country_population.xlsx")

# Updated header for country info
HEADER = [
    "ID",
    "Country",
    "Capital",
    "Region",
    "Subregion",
    "Population",
    "CREATED_AT"
]

def save_to_excel(data: dict) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        next_id = ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.append(HEADER)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center") 

        col_width = [6, 20, 20, 16, 20, 18, 26]
        for i, width in enumerate(col_width, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        next_id = 1

    # Match data to new header
    row = [
        next_id,
        data.get("country", ""),        # Country Name
        data.get("capital", ""),        # Capital City
        data.get("region", ""),         # Region
        data.get("subregion", ""),      # Subregion
        data.get("population", 0),      # Population
        data.get("created_at", "")      # Timestamp
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

if __name__ == "__main__":
    from g3_scrape import scrape_data     # Update import as needed

    data = scrape_data("Asia")               # Example usage; update as needed
    print(data)
    file_path = save_to_excel(data)
    print("Excel file saved to:", file_path)